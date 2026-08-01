from io import BytesIO

from flask import Blueprint, send_file
from sqlalchemy import func

from app import db
from app.models import PointTransaction, Submission, Task, User
from app.services.points import leaderboard
from app.utils.auth import role_required


exports_bp = Blueprint("exports", __name__, url_prefix="/exports")


def points_for(student_id, task_type=None, transaction_type=None, taskless=None):
    query = db.session.query(func.coalesce(func.sum(PointTransaction.points), 0)).filter(
        PointTransaction.student_id == student_id
    )
    if transaction_type:
        query = query.filter(PointTransaction.type == transaction_type)
    if taskless is True:
        query = query.filter(PointTransaction.task_id.is_(None))
    if taskless is False:
        query = query.filter(PointTransaction.task_id.isnot(None))
    if task_type:
        query = query.join(Task, Task.id == PointTransaction.task_id).filter(Task.task_type == task_type)
    return int(query.scalar() or 0)


@exports_bp.route("/excel")
@role_required("admin")
def excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    workbook = Workbook()
    students = User.query.filter_by(role="student").order_by(User.name.asc()).all()
    ranks = {row["name"]: row for row in leaderboard(limit=100000)}
    total_tasks = Task.query.filter_by(status="active").count()

    summary = workbook.active
    summary.title = "Student Summary"
    summary.append(
        [
            "Student",
            "Email",
            "Department",
            "Register No",
            "Daily Points",
            "Weekly Points",
            "Monthly Points",
            "Bonus Points",
            "Penalty Points",
            "Total Points",
            "Completed Tasks",
            "Completion %",
            "Rank",
            "Badge",
        ]
    )

    for student in students:
        daily = points_for(student.id, task_type="daily")
        weekly = points_for(student.id, task_type="weekly")
        monthly = points_for(student.id, task_type="monthly")
        bonus = points_for(student.id, transaction_type="award", taskless=True)
        penalty = points_for(student.id, transaction_type="penalty", taskless=True)
        total = daily + weekly + monthly + bonus + penalty
        completed = Submission.query.filter_by(student_id=student.id, status="approved").count()
        completion = round((completed / total_tasks) * 100, 2) if total_tasks else 0
        rank_data = ranks.get(student.name, {})
        summary.append(
            [
                student.name,
                student.email,
                student.department or "",
                student.register_number or "",
                daily,
                weekly,
                monthly,
                bonus,
                penalty,
                total,
                completed,
                completion,
                rank_data.get("rank", ""),
                rank_data.get("badge", "No Badge"),
            ]
        )

    completion_sheet = workbook.create_sheet("Student Task Completion")
    completion_sheet.append(
        [
            "Student",
            "Email",
            "Department",
            "Task Code",
            "Task Category",
            "Task Stream",
            "Task",
            "Reward Points",
            "Submission Status",
            "Submission Date",
            "Approval Date",
            "Admin Remarks",
        ]
    )
    for student in students:
        submissions = Submission.query.filter_by(student_id=student.id).order_by(Submission.submitted_at.asc()).all()
        if not submissions:
            completion_sheet.append([student.name, student.email, student.department or "", "", "", "", "No submission available", "", "", "", "", ""])
            continue
        for submission in submissions:
            completion_sheet.append(
                [
                    student.name,
                    student.email,
                    student.department or "",
                    submission.task.task_code or "Manual",
                    submission.task.category or "General",
                    submission.task.task_type,
                    submission.task.title,
                    submission.task.reward_points,
                    submission.status,
                    submission.submitted_at,
                    submission.reviewed_at,
                    submission.admin_remarks or "",
                ]
            )

    ledger = workbook.create_sheet("Point Ledger")
    ledger.append(["Student", "Email", "Type", "Points", "Reason", "Task", "Created At"])
    transactions = PointTransaction.query.order_by(PointTransaction.created_at.desc()).all()
    for transaction in transactions:
        ledger.append(
            [
                transaction.student.name,
                transaction.student.email,
                transaction.type,
                transaction.points,
                transaction.reason,
                transaction.task.title if transaction.task else "",
                transaction.created_at,
            ]
        )

    tasks_sheet = workbook.create_sheet("Task Catalogue")
    tasks_sheet.append(["Task Code", "Stream", "Category", "Task", "Reward Points", "Status", "Due Date"])
    for task in Task.query.order_by(Task.task_type.asc(), Task.task_code.asc()).all():
        tasks_sheet.append([task.task_code or "Manual", task.task_type, task.category or "General", task.title, task.reward_points, task.status, task.due_at])

    for sheet in workbook.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="DDEBFF")
        for column_cells in sheet.columns:
            width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 42)
            sheet.column_dimensions[column_cells[0].column_letter].width = width

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name="elite_dashboard_report.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

from datetime import datetime, timedelta
from io import BytesIO

from flask import Blueprint, flash, redirect, render_template, request, send_file, url_for
from flask_login import current_user
from flask_wtf import FlaskForm
from sqlalchemy import func
from wtforms import DateField, SelectField, SubmitField, TimeField
from wtforms.validators import DataRequired

from app import db
from app.models import EliteSprintBid, EliteSprintSession, Task, User
from app.services.elite_sprint import (
    APP_TIMEZONE,
    active_task_ids_by_type,
    get_active_session,
    latest_session as get_latest_session,
    local_datetime_to_utc,
    sprint_leaderboard,
    sprint_metrics,
    utc_datetime_to_local,
    verify_sprint,
)
from app.utils.auth import role_required


elite_sprint_bp = Blueprint("elite_sprint", __name__, url_prefix="/elite-sprint-bidding")


class SprintSessionForm(FlaskForm):
    sprint_type = SelectField(
        "Sprint type",
        choices=[
            ("overall", "Overall Sprint"),
            ("daily", "Daily Sprint"),
            ("weekly", "Weekly Sprint"),
            ("monthly", "Monthly Sprint"),
        ],
        validators=[DataRequired()],
    )
    sprint_date = DateField("Sprint date", validators=[DataRequired()])
    start_time = TimeField("Start time", validators=[DataRequired()])
    end_time = TimeField("End time", validators=[DataRequired()])
    submit = SubmitField("Open Sprint")


@elite_sprint_bp.route("/")
@role_required("admin", "student")
def index():
    if current_user.role == "admin":
        return redirect(url_for("elite_sprint.admin"))
    return redirect(url_for("elite_sprint.student"))


@elite_sprint_bp.route("/admin", methods=["GET", "POST"])
@role_required("admin")
def admin():
    form = SprintSessionForm()
    if form.validate_on_submit():
        local_start_at = datetime.combine(form.sprint_date.data, form.start_time.data)
        local_end_at = datetime.combine(form.sprint_date.data, form.end_time.data)
        if local_end_at <= local_start_at:
            local_end_at = local_end_at + timedelta(days=1)
        start_at = local_datetime_to_utc(local_start_at)
        end_at = local_datetime_to_utc(local_end_at)
        if end_at <= datetime.utcnow():
            flash("End time must be in the future.", "danger")
        else:
            EliteSprintSession.query.filter_by(status="active").update({"status": "closed"})
            session = EliteSprintSession(
                sprint_date=form.sprint_date.data,
                sprint_type=form.sprint_type.data,
                start_time=start_at,
                end_time=end_at,
                status="active",
                created_by=current_user.id,
            )
            db.session.add(session)
            db.session.commit()
            flash("Elite Sprint Bidding is open for the selected window.", "success")
            return redirect(url_for("elite_sprint.admin"))

    active_session = get_active_session()
    latest_session = active_session or get_latest_session()
    return render_template(
        "elite_sprint/admin.html",
        form=form,
        active_session=active_session,
        latest_session=latest_session,
        timezone_label=APP_TIMEZONE,
        latest_start_local=utc_datetime_to_local(latest_session.start_time) if latest_session else None,
        latest_end_local=utc_datetime_to_local(latest_session.end_time) if latest_session else None,
        metrics=sprint_metrics(latest_session),
        leaderboard=sprint_leaderboard(latest_session.id if latest_session else None),
    )


@elite_sprint_bp.route("/admin/verify", methods=["POST"])
@role_required("admin")
def verify_sprint_route():
    session_id = request.form.get("session_id")
    if not session_id:
        flash("No sprint session selected.", "danger")
        return redirect(url_for("elite_sprint.admin"))
    session = EliteSprintSession.query.get_or_404(session_id)
    if session.is_verified:
        flash("Sprint already verified.", "info")
        return redirect(url_for("elite_sprint.admin"))
    try:
        result = verify_sprint(session)
        db.session.commit()
        if result.get("golden_star"):
            flash(f"Sprint verified. Golden Star awarded to {result['golden_star']}.", "success")
        else:
            flash(f"Sprint verified. Penalties applied to {result['penalty_count']} student(s).", "warning")
    except Exception as exc:
        db.session.rollback()
        flash(str(exc), "danger")
    return redirect(url_for("elite_sprint.admin"))


@elite_sprint_bp.route("/admin/close/<int:session_id>", methods=["POST"])
@role_required("admin")
def close(session_id):
    session = EliteSprintSession.query.get_or_404(session_id)
    session.status = "closed"
    db.session.commit()
    leaders = sprint_leaderboard(session.id, limit=1)
    if leaders:
        flash(f"Elite Sprint Bidding has been closed. Highest bidder: {leaders[0]['student_name']} with {leaders[0]['total_tasks']} tasks.", "info")
    else:
        flash("Elite Sprint Bidding has been closed. No bids were submitted.", "info")
    return redirect(url_for("elite_sprint.admin"))


@elite_sprint_bp.route("/student", methods=["GET", "POST"])
@role_required("student")
def student():
    active_session = get_active_session()
    latest_session = active_session or get_latest_session()
    bid = None
    if active_session:
        bid = EliteSprintBid.query.filter_by(session_id=active_session.id, student_id=current_user.id).first()

    if request.method == "POST":
        if not active_session:
            flash("Elite Sprint Closed.", "warning")
            return redirect(url_for("elite_sprint.student"))
        if bid and bid.is_locked:
            flash("Your sprint bid is already locked and cannot be edited.", "danger")
            return redirect(url_for("elite_sprint.student"))
        try:
            daily_tasks = parse_task_ids("daily")
            weekly_tasks = parse_task_ids("weekly")
            monthly_tasks = parse_task_ids("monthly")
            daily_tasks, weekly_tasks, monthly_tasks = filter_tasks_for_sprint_type(
                active_session.sprint_type,
                daily_tasks,
                weekly_tasks,
                monthly_tasks,
            )
            validate_task_ids(daily_tasks, weekly_tasks, monthly_tasks)

            planned_count = len(daily_tasks) + len(weekly_tasks) + len(monthly_tasks)
            task_ids = daily_tasks + weekly_tasks + monthly_tasks

            if planned_count <= 0:
                flash("Enter at least one planned task.", "danger")
                return redirect(url_for("elite_sprint.student"))

            if len(task_ids) != planned_count:
                flash("Number of task IDs must match the planned task count.", "danger")
                return redirect(url_for("elite_sprint.student"))

            if bid is None:
                bid = EliteSprintBid(session_id=active_session.id, student_id=current_user.id)
                db.session.add(bid)
            bid.daily_tasks = daily_tasks
            bid.weekly_tasks = weekly_tasks
            bid.monthly_tasks = monthly_tasks
            bid.daily_count = len(daily_tasks)
            bid.weekly_count = len(weekly_tasks)
            bid.monthly_count = len(monthly_tasks)
            bid.is_locked = True
            bid.locked_at = datetime.utcnow()
            bid.submitted_at = datetime.utcnow()
            db.session.commit()
            flash("Your sprint bid has been saved and locked.", "success")
            return redirect(url_for("elite_sprint.student"))
        except ValueError as exc:
            db.session.rollback()
            flash(str(exc), "danger")

    return render_template(
        "elite_sprint/student.html",
        active_session=active_session,
        latest_session=latest_session,
        bid=bid,
        visible_sections=visible_sections_for_sprint_type(active_session.sprint_type if active_session else "overall"),
        leaderboard=sprint_leaderboard(latest_session.id if latest_session else None),
    )


@elite_sprint_bp.route("/leaderboard")
@role_required("admin", "student")
def leaderboard():
    session = get_active_session() or get_latest_session()
    return render_template(
        "elite_sprint/leaderboard.html",
        session=session,
        leaderboard=sprint_leaderboard(session.id if session else None),
    )


@elite_sprint_bp.route("/analytics")
@role_required("admin")
def analytics():
    session = get_active_session() or get_latest_session()
    metrics = analytics_metrics(session)
    return render_template("elite_sprint/analytics.html", session=session, metrics=metrics)


@elite_sprint_bp.route("/export")
@role_required("admin")
def export():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sprint Bids"
    sheet.append(
        [
            "Session ID",
            "Sprint Date",
            "Sprint Type",
            "Status",
            "Student",
            "Email",
            "Register Number",
            "Department",
            "Daily Count",
            "Daily Tasks",
            "Weekly Count",
            "Weekly Tasks",
            "Monthly Count",
            "Monthly Tasks",
            "Total Sprint Tasks",
            "Submitted At",
        ]
    )
    rows = (
        EliteSprintBid.query.join(EliteSprintSession).join(User)
        .order_by(EliteSprintSession.sprint_date.desc(), EliteSprintBid.submitted_at.asc())
        .all()
    )
    for bid in rows:
        sheet.append(
            [
                bid.session_id,
                bid.session.sprint_date,
                bid.session.sprint_type,
                bid.session.status,
                bid.student.name,
                bid.student.email,
                bid.student.register_number or "",
                bid.student.department or "",
                bid.daily_count,
                ", ".join(map(str, bid.daily_tasks or [])),
                bid.weekly_count,
                ", ".join(map(str, bid.weekly_tasks or [])),
                bid.monthly_count,
                ", ".join(map(str, bid.monthly_tasks or [])),
                bid.total_tasks_bidded,
                bid.submitted_at,
            ]
        )
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDEBFF")
    for column_cells in sheet.columns:
        width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 44)
        sheet.column_dimensions[column_cells[0].column_letter].width = width

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name="elite_sprint_bidding.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def parse_task_ids(section):
    expected_count = parse_non_negative_int(request.form.get(f"{section}_count"), section)
    raw_ids = [value.strip() for value in request.form.getlist(f"{section}_tasks") if value.strip()]
    if len(raw_ids) != expected_count:
        raise ValueError(f"{section.title()} section must contain exactly {expected_count} task ID(s).")
    task_ids = [value.upper() for value in raw_ids]
    if len(task_ids) != len(set(task_ids)):
        raise ValueError(f"Duplicate task IDs are not allowed in the {section.title()} section.")
    return task_ids


def parse_non_negative_int(value, section):
    try:
        number = int(value or 0)
    except ValueError as exc:
        raise ValueError(f"{section.title()} task count must be a number.") from exc
    if number < 0:
        raise ValueError(f"{section.title()} task count cannot be negative.")
    if number > 100:
        raise ValueError(f"{section.title()} task count is too high for one sprint.")
    return number


def validate_task_ids(daily_tasks, weekly_tasks, monthly_tasks):
    grouped, _ = active_task_ids_by_type()
    sections = {
        "Daily": (daily_tasks, grouped["daily"]),
        "Weekly": (weekly_tasks, grouped["weekly"]),
        "Monthly": (monthly_tasks, grouped["monthly"]),
    }
    for label, (task_ids, active_ids) in sections.items():
        invalid = sorted(set(task_ids) - active_ids)
        if invalid:
            raise ValueError(f"{label} task IDs must exist as active {label.lower()} tasks: {', '.join(map(str, invalid))}.")


def visible_sections_for_sprint_type(sprint_type):
    if sprint_type in {"daily", "weekly", "monthly"}:
        return {sprint_type}
    return {"daily", "weekly", "monthly"}


def filter_tasks_for_sprint_type(sprint_type, daily_tasks, weekly_tasks, monthly_tasks):
    visible = visible_sections_for_sprint_type(sprint_type)
    return (
        daily_tasks if "daily" in visible else [],
        weekly_tasks if "weekly" in visible else [],
        monthly_tasks if "monthly" in visible else [],
    )


def analytics_metrics(session):
    if not session:
        return {
            "participants": 0,
            "average_tasks": 0,
            "highest_daily": 0,
            "highest_weekly": 0,
            "highest_monthly": 0,
            "trend_labels": [],
            "trend_values": [],
        }

    totals = db.session.query(
        func.count(EliteSprintBid.id),
        func.coalesce(func.avg(EliteSprintBid.daily_count + EliteSprintBid.weekly_count + EliteSprintBid.monthly_count), 0),
        func.coalesce(func.max(EliteSprintBid.daily_count), 0),
        func.coalesce(func.max(EliteSprintBid.weekly_count), 0),
        func.coalesce(func.max(EliteSprintBid.monthly_count), 0),
    ).filter(EliteSprintBid.session_id == session.id).one()
    trend_rows = (
        db.session.query(EliteSprintSession.sprint_date, func.count(EliteSprintBid.id))
        .outerjoin(EliteSprintBid, EliteSprintBid.session_id == EliteSprintSession.id)
        .group_by(EliteSprintSession.sprint_date)
        .order_by(EliteSprintSession.sprint_date.asc())
        .limit(30)
        .all()
    )
    return {
        "participants": int(totals[0] or 0),
        "average_tasks": round(float(totals[1] or 0), 2),
        "highest_daily": int(totals[2] or 0),
        "highest_weekly": int(totals[3] or 0),
        "highest_monthly": int(totals[4] or 0),
        "trend_labels": [row[0].strftime("%d %b") for row in trend_rows],
        "trend_values": [int(row[1] or 0) for row in trend_rows],
    }

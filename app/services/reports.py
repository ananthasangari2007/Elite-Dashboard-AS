import os
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app import db
from app.models import EliteSprintBid, EliteSprintSession, PointTransaction, Submission, Task, User

REPORTS_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "reports")


def ensure_reports_dir():
    os.makedirs(REPORTS_DIR, exist_ok=True)


def month_workbook_path(year, month):
    ensure_reports_dir()
    month_name = datetime(year, month, 1).strftime("%B").upper()
    return os.path.join(REPORTS_DIR, f"{month_name}-{year} TASK TRACK SHEET.xlsx")


def date_sheet_name(value):
    return value.strftime("%d.%m.%Y")


def _apply_sheet_format(sheet, header_fill="1F4E79"):
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=header_fill)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = Border(
                left=Side(style="thin", color="CCCCCC"),
                right=Side(style="thin", color="CCCCCC"),
                top=Side(style="thin", color="CCCCCC"),
                bottom=Side(style="thin", color="CCCCCC"),
            )
            if cell.row > 1:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
    for column_cells in sheet.columns:
        width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 42)
        sheet.column_dimensions[column_cells[0].column_letter].width = width


def _status_symbol(status):
    if status == "approved":
        return "✅"
    if status == "rejected":
        return "❌"
    if status == "waiting_approval":
        return "⚠️"
    return "-"


def _get_task_display(task_id, task_map):
    task = task_map.get(task_id)
    if task:
        return task.task_code or str(task.id)
    return str(task_id)


def _get_bids_for_date(sprint_date):
    session = (
        EliteSprintSession.query.filter_by(sprint_date=sprint_date)
        .order_by(EliteSprintSession.created_at.desc())
        .first()
    )
    if not session:
        return [], None, {}
    bids = (
        EliteSprintBid.query.filter_by(session_id=session.id)
        .join(User, User.id == EliteSprintBid.student_id)
        .order_by(User.name.asc())
        .all()
    )
    task_ids = set()
    for bid in bids:
        task_ids.update(bid.daily_tasks or [])
        task_ids.update(bid.weekly_tasks or [])
        task_ids.update(bid.monthly_tasks or [])
    tasks = Task.query.filter(Task.id.in_(list(task_ids))).all() if task_ids else []
    task_map = {str(task.id): task for task in tasks}
    return bids, session, task_map


def _get_all_active_tasks():
    tasks = Task.query.filter_by(status="active").order_by(Task.task_type.asc(), Task.created_at.asc()).all()
    task_map = {str(task.id): task for task in tasks}
    daily_ids = sorted({str(task.id) for task in tasks if task.task_type == "daily"})
    weekly_ids = sorted({str(task.id) for task in tasks if task.task_type == "weekly"})
    monthly_ids = sorted({str(task.id) for task in tasks if task.task_type == "monthly"})
    return daily_ids, weekly_ids, monthly_ids, task_map


def _get_task_status_for_student(student_id, sprint_date, task_map, daily_ids, weekly_ids, monthly_ids):
    submissions = (
        Submission.query.filter_by(student_id=student_id)
        .filter(Submission.submission_date == sprint_date)
        .all()
    )
    submission_map = {}
    for sub in submissions:
        task = sub.task
        key = str(task.task_code.strip().upper() if task.task_code else task.id)
        submission_map[key] = sub.status

    bid = EliteSprintBid.query.filter_by(student_id=student_id).join(EliteSprintSession).filter(
        EliteSprintSession.sprint_date == sprint_date
    ).first()
    bidded_ids = set()
    if bid:
        bidded_ids.update(str(t).upper() for t in (bid.daily_tasks or []))
        bidded_ids.update(str(t).upper() for t in (bid.weekly_tasks or []))
        bidded_ids.update(str(t).upper() for t in (bid.monthly_tasks or []))

    def cell_status(task_id):
        display = _get_task_display(task_id, task_map).upper()
        if display in submission_map:
            return _status_symbol(submission_map[display])
        if display in bidded_ids:
            return "⚠️"
        return "-"

    return daily_ids, weekly_ids, monthly_ids, cell_status


def _points_for_student(student_id, task_type=None, transaction_type=None, taskless=None, sprint_date=None):
    query = db.session.query(db.func.coalesce(db.func.sum(PointTransaction.points), 0)).filter(
        PointTransaction.student_id == student_id
    )
    if transaction_type:
        query = query.filter(PointTransaction.type == transaction_type)
    if taskless is True:
        query = query.filter(PointTransaction.task_id.is_(None))
    if taskless is False:
        query = query.filter(PointTransaction.task_id.isnot(None))
    if task_type:
        query = query.join(Task, Task.id == PointTransaction.task_id).filter(Task.task_type == task_type)
    if sprint_date:
        start = datetime.combine(sprint_date, datetime.min.time())
        end = datetime.combine(sprint_date, datetime.max.time())
        query = query.filter(PointTransaction.created_at >= start, PointTransaction.created_at <= end)
    return int(query.scalar() or 0)


def build_bidding_sheet(workbook, sheet_name, sprint_date):
    if sheet_name in workbook.sheetnames:
        idx = workbook.sheetnames.index(sheet_name)
        workbook.remove(workbook.worksheets[idx])
    sheet = workbook.create_sheet(sheet_name)
    headers = [
        "STUDENT NAME",
        "EMAIL ID",
        "NO OF DAILY TASK PLANNED",
        "LIST OF DAILY TASK ID PLANNED",
        "NO OF WEEKLY TASK PLANNED",
        "LIST OF WEEKLY TASK ID PLANNED",
        "NO OF MONTHLY TASK PLANNED",
        "LIST OF MONTHLY TASK ID PLANNED",
    ]
    sheet.append(headers)

    bids, _, task_map = _get_bids_for_date(sprint_date)
    for bid in bids:
        daily = sorted(bid.daily_tasks or [])
        weekly = sorted(bid.weekly_tasks or [])
        monthly = sorted(bid.monthly_tasks or [])
        sheet.append(
            [
                bid.student.name.upper(),
                bid.student.email,
                len(daily),
                ", ".join(_get_task_display(str(t), task_map) for t in daily) if daily else "-",
                len(weekly),
                ", ".join(_get_task_display(str(t), task_map) for t in weekly) if weekly else "-",
                len(monthly),
                ", ".join(_get_task_display(str(t), task_map) for t in monthly) if monthly else "-",
            ]
        )

    _apply_sheet_format(sheet)
    return sheet


def build_task_status_sheet(workbook, sheet_name, sprint_date):
    if sheet_name in workbook.sheetnames:
        idx = workbook.sheetnames.index(sheet_name)
        workbook.remove(workbook.worksheets[idx])
    sheet = workbook.create_sheet(sheet_name)

    daily_ids, weekly_ids, monthly_ids, task_map = _get_all_active_tasks()
    headers = ["STUDENT NAME", "EMAIL ID"]
    headers.extend(daily_ids)
    headers.extend(weekly_ids)
    headers.extend(monthly_ids)
    headers.extend(
        [
            "BONUS POINTS",
            "PENALTY POINTS",
            "TODAY POINTS",
            "PREVIOUS DAY BALANCE",
            "MONTHLY TOTAL POINTS",
        ]
    )
    sheet.append(headers)

    start = datetime.combine(sprint_date, datetime.min.time())
    end = datetime.combine(sprint_date, datetime.max.time())

    month_start = datetime(sprint_date.year, sprint_date.month, 1)
    if sprint_date.month == 12:
        month_end = datetime(sprint_date.year + 1, 1, 1)
    else:
        month_end = datetime(sprint_date.year, sprint_date.month + 1, 1)

    students = User.query.filter_by(role="student").order_by(User.name.asc()).all()
    for student in students:
        _, _, _, status_fn = _get_task_status_for_student(student.id, sprint_date, task_map, daily_ids, weekly_ids, monthly_ids)
        row = [student.name, student.email]
        row.extend(status_fn(tid) for tid in daily_ids)
        row.extend(status_fn(tid) for tid in weekly_ids)
        row.extend(status_fn(tid) for tid in monthly_ids)

        bonus = (
            db.session.query(db.func.coalesce(db.func.sum(PointTransaction.points), 0))
            .filter(
                PointTransaction.student_id == student.id,
                PointTransaction.type == "award",
                PointTransaction.task_id.is_(None),
                PointTransaction.created_at >= start,
                PointTransaction.created_at <= end,
            )
            .scalar()
            or 0
        )
        penalty = (
            db.session.query(db.func.coalesce(db.func.sum(PointTransaction.points), 0))
            .filter(
                PointTransaction.student_id == student.id,
                PointTransaction.type == "penalty",
                PointTransaction.task_id.is_(None),
                PointTransaction.created_at >= start,
                PointTransaction.created_at <= end,
            )
            .scalar()
            or 0
        )

        today_points = (
            db.session.query(db.func.coalesce(db.func.sum(PointTransaction.points), 0))
            .filter(
                PointTransaction.student_id == student.id,
                PointTransaction.created_at >= start,
                PointTransaction.created_at <= end,
            )
            .scalar()
            or 0
        )

        if sprint_date.day > 1:
            prev_date = sprint_date.replace(day=sprint_date.day - 1)
        else:
            prev_date = sprint_date
        prev_start = datetime.combine(prev_date, datetime.min.time())
        prev_end = datetime.combine(sprint_date, datetime.min.time())
        prev_day_points = (
            db.session.query(db.func.coalesce(db.func.sum(PointTransaction.points), 0))
            .filter(
                PointTransaction.student_id == student.id,
                PointTransaction.created_at >= prev_start,
                PointTransaction.created_at < prev_end,
            )
            .scalar()
            or 0
        )

        monthly_total = (
            db.session.query(db.func.coalesce(db.func.sum(PointTransaction.points), 0))
            .filter(
                PointTransaction.student_id == student.id,
                PointTransaction.created_at >= month_start,
                PointTransaction.created_at < month_end,
            )
            .scalar()
            or 0
        )

        row.extend([int(bonus), int(penalty), int(today_points), int(prev_day_points), int(monthly_total)])
        sheet.append(row)

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            if cell.value == "✅":
                cell.fill = PatternFill("solid", fgColor="C6EFCE")
                cell.font = Font(color="006100", bold=True)
            elif cell.value == "❌":
                cell.fill = PatternFill("solid", fgColor="FFC7CE")
                cell.font = Font(color="9C0006", bold=True)
            elif cell.value == "⚠️":
                cell.fill = PatternFill("solid", fgColor="FFEB9C")
                cell.font = Font(color="9C6500", bold=True)
            elif cell.value == "⭐":
                cell.fill = PatternFill("solid", fgColor="FFD966")
                cell.font = Font(color="7F6000", bold=True)

    _apply_sheet_format(sheet)
    return sheet


def build_daily_summary_sheet(workbook, sheet_name, sprint_date):
    if sheet_name in workbook.sheetnames:
        idx = workbook.sheetnames.index(sheet_name)
        workbook.remove(workbook.worksheets[idx])
    sheet = workbook.create_sheet(sheet_name)

    students = User.query.filter_by(role="student").all()
    bids, session, task_map = _get_bids_for_date(sprint_date)
    start = datetime.combine(sprint_date, datetime.min.time())
    end = datetime.combine(sprint_date, datetime.max.time())

    total_students = len(students)
    participants = len({bid.student_id for bid in bids})
    daily_planned = sum(bid.daily_count for bid in bids)
    weekly_planned = sum(bid.weekly_count for bid in bids)
    monthly_planned = sum(bid.monthly_count for bid in bids)

    approved_count = Submission.query.filter_by(submission_date=sprint_date, status="approved").count()
    rejected_count = Submission.query.filter_by(submission_date=sprint_date, status="rejected").count()
    pending_count = Submission.query.filter_by(submission_date=sprint_date, status="waiting_approval").count()

    bonus_points = (
        db.session.query(db.func.coalesce(db.func.sum(PointTransaction.points), 0))
        .filter(
            PointTransaction.type == "award",
            PointTransaction.task_id.is_(None),
            PointTransaction.created_at >= start,
            PointTransaction.created_at <= end,
        )
        .scalar()
        or 0
    )
    penalty_points = (
        db.session.query(db.func.coalesce(db.func.sum(PointTransaction.points), 0))
        .filter(
            PointTransaction.type == "penalty",
            PointTransaction.task_id.is_(None),
            PointTransaction.created_at >= start,
            PointTransaction.created_at <= end,
        )
        .scalar()
        or 0
    )

    golden_stars = EliteSprintBid.query.join(EliteSprintSession).filter(
        EliteSprintSession.sprint_date == sprint_date,
        EliteSprintBid.has_golden_star.is_(True),
    ).count()

    penalty_cases = (
        db.session.query(db.func.count(PointTransaction.id))
        .filter(
            PointTransaction.type == "penalty",
            PointTransaction.created_at >= start,
            PointTransaction.created_at <= end,
        )
        .scalar()
        or 0
    )

    today_points = []
    for student in students:
        pts = (
            db.session.query(db.func.coalesce(db.func.sum(PointTransaction.points), 0))
            .filter(
                PointTransaction.student_id == student.id,
                PointTransaction.created_at >= start,
                PointTransaction.created_at <= end,
            )
            .scalar()
            or 0
        )
        today_points.append((student.name, int(pts)))

    top_performer = max(today_points, key=lambda x: x[1])[0] if today_points else "N/A"

    compliance = 0.0
    if participants:
        golden_star_count = EliteSprintBid.query.join(EliteSprintSession).filter(
            EliteSprintSession.sprint_date == sprint_date,
            EliteSprintBid.has_golden_star.is_(True),
        ).count()
        compliance = round((golden_star_count / participants) * 100, 2)

    summary = [
        ["TOTAL STUDENTS", total_students],
        ["STUDENTS PARTICIPATED IN BIDDING", participants],
        ["TOTAL DAILY TASKS PLANNED", daily_planned],
        ["TOTAL TASKS APPROVED", approved_count],
        ["TOTAL TASKS REJECTED", rejected_count],
        ["TOTAL BONUS POINTS", int(bonus_points)],
        ["TOTAL PENALTY POINTS", int(penalty_points)],
        ["GOLDEN STARS AWARDED", golden_stars],
        ["PENALTY CASES", penalty_cases],
        ["TOP PERFORMER OF THE DAY", top_performer],
        ["SPRINT COMPLIANCE RATE", f"{compliance}%"],
    ]

    sheet.append(["METRIC", "VALUE"])
    for row in summary:
        sheet.append(row)

    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in sheet.iter_rows(min_row=2):
        row[0].font = Font(bold=True)
        row[0].fill = PatternFill("solid", fgColor="DDEBFF")
        row[0].alignment = Alignment(vertical="center")
        row[1].alignment = Alignment(vertical="center")
        for cell in row:
            cell.border = Border(
                left=Side(style="thin", color="CCCCCC"),
                right=Side(style="thin", color="CCCCCC"),
                top=Side(style="thin", color="CCCCCC"),
                bottom=Side(style="thin", color="CCCCCC"),
            )

    chart_start_row = len(summary) + 3
    sheet.cell(row=chart_start_row, column=1, value="CATEGORY")
    sheet.cell(row=chart_start_row, column=2, value="COUNT")
    sheet.append(["Approved", approved_count])
    sheet.append(["Rejected", rejected_count])
    sheet.append(["Pending", pending_count])

    pie = PieChart()
    pie.title = "Approved vs Rejected vs Pending"
    pie.height = 7
    pie.width = 12
    labels = Reference(sheet, min_col=1, min_row=chart_start_row + 1, max_row=chart_start_row + 3)
    data = Reference(sheet, min_col=2, min_row=chart_start_row, max_row=chart_start_row + 3)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    sheet.add_chart(pie, f"D{chart_start_row}")

    top5 = sorted(today_points, key=lambda x: x[1], reverse=True)[:5]
    top5_start = chart_start_row + 6
    sheet.cell(row=top5_start, column=1, value="STUDENT")
    sheet.cell(row=top5_start, column=2, value="TODAY POINTS")
    for idx, (name, pts) in enumerate(top5, start=1):
        sheet.cell(row=top5_start + idx, column=1, value=name)
        sheet.cell(row=top5_start + idx, column=2, value=pts)

    bar = BarChart()
    bar.type = "bar"
    bar.title = "Top 5 students by Today Points"
    bar.height = 7
    bar.width = 12
    labels = Reference(sheet, min_col=1, min_row=top5_start + 1, max_row=top5_start + 5)
    data = Reference(sheet, min_col=2, min_row=top5_start, max_row=top5_start + 5)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(labels)
    sheet.add_chart(bar, f"D{top5_start}")

    column_start = top5_start + 8
    sheet.cell(row=column_start, column=1, value="TASK TYPE")
    sheet.cell(row=column_start, column=2, value="PLANNED")
    sheet.append(["Daily", daily_planned])
    sheet.append(["Weekly", weekly_planned])
    sheet.append(["Monthly", monthly_planned])

    column = BarChart()
    column.type = "col"
    column.title = "Daily vs Weekly vs Monthly planned task counts"
    column.height = 7
    column.width = 12
    labels = Reference(sheet, min_col=1, min_row=column_start + 1, max_row=column_start + 3)
    data = Reference(sheet, min_col=2, min_row=column_start, max_row=column_start + 3)
    column.add_data(data, titles_from_data=True)
    column.set_categories(labels)
    sheet.add_chart(column, f"D{column_start}")

    for column_cells in sheet.columns:
        width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 42)
        sheet.column_dimensions[column_cells[0].column_letter].width = width

    return sheet


def export_daily_report(sprint_date=None):
    sprint_date = sprint_date or datetime.utcnow().date()
    year = sprint_date.year
    month = sprint_date.month
    path = month_workbook_path(year, month)
    sheet_prefix = date_sheet_name(sprint_date)

    if os.path.exists(path):
        workbook = load_workbook(path)
    else:
        from openpyxl import Workbook
        workbook = Workbook()
        if "Sheet" in workbook.sheetnames:
            workbook.remove(workbook["Sheet"])

    build_bidding_sheet(workbook, f"{sheet_prefix} - BIDDING SHEET", sprint_date)
    build_task_status_sheet(workbook, f"{sheet_prefix} - TASK STATUS", sprint_date)
    build_daily_summary_sheet(workbook, f"{sheet_prefix} - DAILY SUMMARY", sprint_date)

    workbook.save(path)
    return path
